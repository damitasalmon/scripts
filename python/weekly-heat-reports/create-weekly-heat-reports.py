from openpyxl import Workbook
from datetime import datetime


# Get today's date
today = datetime.today()


# Format the date as MMDDYY
today_date_str = today.strftime('%m%d%y')

#Update the workbook object, worksheet and file names as needed

# Create New workbook and add sheets
wb_in = Workbook()
ws_in = wb_in.active
ws_in.title = "New sheet title"
in_ws2 = wb_in.create_sheet(title="Sheet2")
in_ws3 = wb_in.create_sheet(title="Sheet3")
in_ws4 = wb_in.create_sheet(title="UKG")
in_ws5 = wb_in.create_sheet(title="CHARM")

# Save the workbook
wb_in.save(f"C:/Users/.../..../...Report AS OF {today_date_str}.xlsx")

# Create New workbook and add sheets
wb_sr = Workbook()
ws_sr = wb_sr.active
ws_sr.title = "New sheet title"
sr_ws2 = wb_sr.create_sheet(title="Sheet2")
sr_ws3 = wb_sr.create_sheet(title="Sheet3")

# Save the workbook
wb_sr.save(f"C:/Users/.../..../...Report AS OF {today_date_str}.xlsx")

# Create New workbook and add sheets
wb_tsk = Workbook()
ws_tsk = wb_tsk.active
ws_tsk.title = "New sheet title"
ws2 = wb_tsk.create_sheet(title="Sheet2")
ws3 = wb_tsk.create_sheet(title="Sheet3")

# Save the workbook
wb_tsk.save(f"C:/Users/.../..../...Report AS OF {today_date_str}.xlsx")